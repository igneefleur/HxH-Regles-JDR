# -*- coding: utf-8 -*-
"""
Construit une fiche de personnage générique, entièrement précalculée, au format .xlsx
pour le JDR maison Hunter x Hunter.

Trois feuilles :
  - "Fiche"     : la fiche jouable (saisies en blanc, calculs en parchemin).
  - "Barèmes"   : toutes les tables de correspondance (masquée). Référencée par la Fiche.
  - "Référence" : aide-mémoire consultatif (difficulté, dégâts, états, manœuvres, armes).

Formules écrites avec les noms ANGLAIS (VLOOKUP, IF, INDEX, MATCH, FLOOR, CEILING, ...) :
Excel et Google Sheets les convertissent automatiquement vers la locale d'ouverture.

Lancer :  python docs/assets/build_fiche.py
Sortie :  docs/assets/fiche-personnage-hxh.xlsx
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# --------------------------------------------------------------------------- #
# Couleurs (thème du livre)
# --------------------------------------------------------------------------- #
C_PARCHEMIN = "FAF6EB"  # fond parchemin (cellules calculées)
C_TITRE     = "667861"  # vert titre
C_SECTION   = "63755E"  # vert section (titres de bloc)
C_ENTETE    = "82997B"  # vert en-tête (en-têtes de tableau)
C_ALT1      = "D6E2D0"  # ligne alternée 1
C_ALT2      = "E9F0E6"  # ligne alternée 2
C_ENCRE     = "1D1D1D"  # encre
C_BLANC     = "FFFFFF"  # cellules de saisie
C_LEGENDE   = "F2E9CE"  # léger fond pour la légende

# --------------------------------------------------------------------------- #
# Styles réutilisables
# --------------------------------------------------------------------------- #
FONT_BASE    = Font(name="Calibri", size=10, color=C_ENCRE)
FONT_CALC    = Font(name="Calibri", size=10, color=C_ENCRE)
FONT_INPUT   = Font(name="Calibri", size=10, color="1D1D1D")
FONT_LABEL   = Font(name="Calibri", size=10, bold=True, color=C_ENCRE)
FONT_SECTION = Font(name="Calibri", size=12, bold=True, color=C_BLANC)
FONT_TITRE   = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
FONT_ENTETE  = Font(name="Calibri", size=10, bold=True, color=C_BLANC)
FONT_SMALL   = Font(name="Calibri", size=8, italic=True, color="555555")

FILL_PARCHEMIN = PatternFill("solid", fgColor=C_PARCHEMIN)
FILL_SECTION   = PatternFill("solid", fgColor=C_SECTION)
FILL_TITRE     = PatternFill("solid", fgColor=C_TITRE)
FILL_ENTETE    = PatternFill("solid", fgColor=C_ENTETE)
FILL_INPUT     = PatternFill("solid", fgColor=C_BLANC)
FILL_ALT1      = PatternFill("solid", fgColor=C_ALT1)
FILL_ALT2      = PatternFill("solid", fgColor=C_ALT2)
FILL_LEGENDE   = PatternFill("solid", fgColor=C_LEGENDE)

_thin = Side(style="thin", color="9AA89A")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

AL_L = Alignment(horizontal="left",   vertical="center", wrap_text=False)
AL_C = Alignment(horizontal="center", vertical="center", wrap_text=False)
AL_R = Alignment(horizontal="right",  vertical="center", wrap_text=False)
AL_LW = Alignment(horizontal="left",  vertical="center", wrap_text=True)
AL_CW = Alignment(horizontal="center", vertical="center", wrap_text=True)


# --------------------------------------------------------------------------- #
# Helpers d'écriture
# --------------------------------------------------------------------------- #
def cell(ws, coord, value=None, *, font=None, fill=None, border=True,
         align=None, fmt=None):
    """Écrit une cellule avec style."""
    c = ws[coord]
    if value is not None:
        c.value = value
    c.font = font or FONT_BASE
    if fill is not None:
        c.fill = fill
    if border:
        c.border = BORDER
    if align is not None:
        c.alignment = align
    if fmt is not None:
        c.number_format = fmt
    return c


def label(ws, coord, text, **kw):
    return cell(ws, coord, text, font=FONT_LABEL, fill=FILL_PARCHEMIN,
               align=AL_L, **kw)


def calc(ws, coord, formula, **kw):
    """Cellule calculée (formule) : fond parchemin."""
    kw.setdefault("align", AL_C)
    return cell(ws, coord, formula, font=FONT_CALC, fill=FILL_PARCHEMIN, **kw)


def inp(ws, coord, value=None, **kw):
    """Cellule de saisie : fond blanc."""
    kw.setdefault("align", AL_C)
    return cell(ws, coord, value, font=FONT_INPUT, fill=FILL_INPUT, **kw)


def section_title(ws, row, col_start, col_end, text):
    """Bandeau de titre de section vert, fusionné."""
    a = f"{get_column_letter(col_start)}{row}"
    b = f"{get_column_letter(col_end)}{row}"
    ws.merge_cells(f"{a}:{b}")
    cell(ws, a, text, font=FONT_SECTION, fill=FILL_SECTION, align=AL_L)
    for cc in range(col_start, col_end + 1):
        ws[f"{get_column_letter(cc)}{row}"].border = BORDER
        ws[f"{get_column_letter(cc)}{row}"].fill = FILL_SECTION
    ws.row_dimensions[row].height = 20


def header_cell(ws, coord, text):
    return cell(ws, coord, text, font=FONT_ENTETE, fill=FILL_ENTETE, align=AL_CW)


# =========================================================================== #
# DONNÉES (valeurs exactes du JSON)
# =========================================================================== #

# --- Modificateur par caractéristique (0..30) ------------------------------ #
def mod_for_value(v):
    if v == 0:  return -25
    if v == 1:  return -20
    if v == 2:  return -15
    if v == 3:  return -10
    if v == 4:  return -5
    if v == 5:  return 0
    if 6 <= v <= 9:   return 5
    if 10 <= v <= 12: return 10
    if 13 <= v <= 14: return 15
    if 15 <= v <= 17: return 20
    if 18 <= v <= 19: return 25
    if 20 <= v <= 22: return 30
    if 23 <= v <= 24: return 35
    if 25 <= v <= 27: return 40
    if 28 <= v <= 29: return 45
    if v == 30: return 50
    return 0

MOD_TABLE = [(v, mod_for_value(v)) for v in range(31)]

# --- Plafond carac par Éclat ----------------------------------------------- #
# Éclat (palier) -> plafond de carac
PLAFOND_ECLAT = [(0, 9), (10, 14), (20, 19), (30, 24), (40, 29), (50, 30)]

# --- Création par Éclat : moyenne / plafond / budget ----------------------- #
# borne basse d'Éclat -> (moyenne, plafond, budget)
CREATION_ECLAT = [(0, 5, 7, 24), (15, 7, 9, 48), (30, 9, 11, 72), (45, 11, 13, 96)]

# --- Indice de poids par Force (texte exact) ------------------------------- #
POIDS = {
    0:  ("<1 kg", "1 kg", "2 kg"),
    1:  ("1 kg", "2 kg", "5 kg"),
    2:  ("2 kg", "5 kg", "10 kg"),
    3:  ("3 kg", "8 kg", "20 kg"),
    4:  ("5 kg", "15 kg", "50 kg"),
    5:  ("10 kg", "30 kg", "100 kg"),
    6:  ("20 kg", "60 kg", "200 kg"),
    7:  ("30 kg", "100 kg", "300 kg"),
    8:  ("50 kg", "150 kg", "500 kg"),
    9:  ("100 kg", "300 kg", "1 t"),
    10: ("1 t", "3 t", "10 t"),
    11: ("2 t", "6 t", "20 t"),
    12: ("3 t", "10 t", "30 t"),
    13: ("5 t", "15 t", "50 t"),
    14: ("10 t", "30 t", "100 t"),
    15: ("100 t", "300 t", "1 000 t"),
    16: ("200 t", "600 t", "2 000 t"),
    17: ("300 t", "1 000 t", "3 000 t"),
    18: ("500 t", "1 500 t", "5 000 t"),
    19: ("1 000 t", "3 000 t", "10 000 t"),
    20: ("10 000 t", "30 000 t", "100 000 t"),
    21: ("20 000 t", "60 000 t", "200 000 t"),
    22: ("30 000 t", "100 000 t", "300 000 t"),
    23: ("50 000 t", "150 000 t", "500 000 t"),
    24: ("100 000 t", "300 000 t", "1 000 000 t"),
    25: ("1 000 000 t", "3 000 000 t", "10 000 000 t"),
    26: ("2M t", "6M t", "20M t"),
    27: ("3M t", "10M t", "30M t"),
    28: ("5M t", "15M t", "50M t"),
    29: ("10M t", "30M t", "100M t"),
    30: ("100M t", "300M t", "1 000M t"),
}

# --- Mouvement par Agilité (par round) ------------------------------------- #
MOUVEMENT = {
    0:  ("1 m", "2 m", "5 m"),
    1:  ("2 m", "4 m", "10 m"),
    2:  ("3 m", "6 m", "15 m"),
    3:  ("4 m", "8 m", "20 m"),
    4:  ("6 m", "12 m", "30 m"),
    5:  ("8 m", "16 m", "40 m"),
    6:  ("10 m", "20 m", "50 m"),
    7:  ("12 m", "24 m", "60 m"),
    8:  ("14 m", "28 m", "70 m"),
    9:  ("16 m", "32 m", "80 m"),
    10: ("20 m", "40 m", "100 m"),
    11: ("30 m", "60 m", "150 m"),
    12: ("40 m", "80 m", "200 m"),
    13: ("60 m", "120 m", "300 m"),
    14: ("80 m", "160 m", "400 m"),
    15: ("100 m", "200 m", "500 m"),
    16: ("150 m", "300 m", "750 m"),
    17: ("200 m", "400 m", "1 km"),
    18: ("300 m", "600 m", "1,5 km"),
    19: ("400 m", "800 m", "2 km"),
    20: ("600 m", "1,2 km", "3 km"),
    21: ("800 m", "1,6 km", "4 km"),
    22: ("1 km", "2 km", "5 km"),
    23: ("1,5 km", "3 km", "7,5 km"),
    24: ("2 km", "4 km", "10 km"),
    25: ("3 km", "6 km", "15 km"),
    26: ("5 km", "10 km", "25 km"),
    27: ("10 km", "20 km", "50 km"),
    28: ("20 km", "40 km", "100 km"),
    29: ("50 km", "100 km", "250 km"),
    30: ("100 km", "200 km", "500 km"),
}

# --- Apnée par Endurance --------------------------------------------------- #
APNEE = {
    0:  ("<6 s", "0 s", "0 s"),
    1:  ("6 s", "<6 s", "0 s"),
    2:  ("12 s", "6 s", "<6 s"),
    3:  ("18 s", "12 s", "6 s"),
    4:  ("30 s", "18 s", "12 s"),
    5:  ("1 min", "30 s", "18 s"),
    6:  ("2 min", "1 min", "30 s"),
    7:  ("3 min", "2 min", "1 min"),
    8:  ("5 min", "3 min", "2 min"),
    9:  ("10 min", "5 min", "3 min"),
    10: ("20 min", "10 min", "5 min"),
    11: ("30 min", "20 min", "10 min"),
    12: ("1 h", "30 min", "20 min"),
    13: ("2 h", "1 h", "30 min"),
    14: ("3 h", "2 h", "1 h"),
    15: ("5 h", "3 h", "2 h"),
    16: ("10 h", "5 h", "3 h"),
    17: ("1 jour", "10 h", "5 h"),
    18: ("2 jours", "1 jour", "10 h"),
    19: ("3 jours", "2 jours", "1 jour"),
    20: ("5 jours", "3 jours", "2 jours"),
    21: ("10 jours", "5 jours", "3 jours"),
    22: ("20 jours", "10 jours", "5 jours"),
    23: ("1 mois", "20 jours", "10 jours"),
    24: ("2 mois", "1 mois", "20 jours"),
    25: ("3 mois", "2 mois", "1 mois"),
    26: ("5 mois", "3 mois", "2 mois"),
    27: ("10 mois", "5 mois", "3 mois"),
    28: ("illimitée", "10 mois", "5 mois"),
    29: ("illimitée", "illimitée", "10 mois"),
    30: ("illimitée", "illimitée", "illimitée"),
}

# --- Affinités par archétype ----------------------------------------------- #
# nom -> (Ren, Émi, Tra, Man, Conj, Spé_sans=0, Spé_avec)
ARCHETYPES = [
    ("Renforceur",             100, 80, 80, 60, 60, 0, 40),
    ("Émitteur",               80, 100, 60, 80, 40, 0, 60),
    ("Transmuteur",            80, 60, 100, 40, 80, 0, 60),
    ("Manipulateur",           60, 80, 40, 100, 60, 0, 80),
    ("Conjurateur",            60, 40, 80, 60, 100, 0, 80),
    ("Spécialiste",            40, 60, 60, 80, 80, 0, 100),
    ("Renforceur-Émitteur",    90, 90, 70, 70, 50, 0, 50),
    ("Renforceur-Transmuteur", 90, 70, 90, 50, 70, 0, 50),
    ("Émitteur-Manipulateur",  70, 90, 50, 90, 50, 0, 70),
    ("Transmuteur-Conjurateur",70, 50, 90, 50, 90, 0, 70),
    ("Manipulateur-Spécialiste",50, 70, 50, 90, 70, 0, 90),
    ("Conjurateur-Spécialiste", 50, 50, 70, 70, 90, 0, 90),
]

# --- Coûts DI des techniques par palier ------------------------------------ #
# technique -> dict palier->coût cumulé (pour les techniques à paliers)
# Pour les blocs (Initiation/Hatsu/Gyo/Shu/In) : coût unique.
TECH_PALIERS = {  # cumulés Aucun/Basique/Avancé/Expert/Maître
    "Ten":   {"Aucun": 0, "Basique": 20, "Avancé": 30, "Expert": 50, "Maître": 80},
    "Zetsu": {"Aucun": 0, "Basique": 10, "Avancé": 20, "Expert": 40, "Maître": 70},
    "Ren":   {"Aucun": 0, "Basique": 20, "Avancé": 30, "Expert": 50, "Maître": 80},
    "Ken":   {"Aucun": 0, "Basique": 20, "Avancé": 30, "Expert": 50, "Maître": 80},
    "En":    {"Aucun": 0, "Basique": 20, "Avancé": 30, "Expert": 50, "Maître": 80},
    "Ko":    {"Aucun": 0, "Basique": 40, "Avancé": 50, "Expert": 70, "Maître": 100},
    "Ryu":   {"Aucun": 0, "Basique": 40, "Avancé": 50, "Expert": 70, "Maître": 100},
}
TECH_BLOCS = {  # technique -> coût DI (case Oui/Non)
    "Initiation": 40, "Hatsu": 30, "Gyo": 20, "Shu": 20, "In": 30,
}
# Base d'UAR conférée par l'état à manteau actif (base 5 + augmentation du palier)
UAR_PAR_PALIER = {"Aucun": 5, "Basique": 10, "Avancé": 15, "Expert": 20, "Maître": 25}

# --- Échelle de difficulté (13 paliers) ------------------------------------ #
DIFFICULTE = [
    ("Triviale", 0), ("Très facile", 20), ("Facile", 40), ("Moyenne", 80),
    ("Difficile", 120), ("Très difficile", 180), ("Absurde", 240),
    ("Quasi impossible", 320), ("Impossible", 400), ("Surhumaine", 520),
    ("Prodigieuse", 640), ("Insurmontable", 780), ("Inimaginable", 920),
]

# --- Types de dégâts (7) --------------------------------------------------- #
TYPES_DEGATS = [
    ("CON", "Contondant", "Physique", "Masses, poings, chutes"),
    ("TRA", "Tranchant", "Physique", "Lames, griffes, fil de Nen"),
    ("PER", "Perforant", "Physique", "Pointes, flèches, balles"),
    ("FEU", "Feu", "Élémentaire", "Feu, lave, vapeur"),
    ("FRO", "Froid", "Élémentaire", "Glace, gel"),
    ("ÉLE", "Électrique", "Élémentaire", "Foudre, décharges, lumière"),
    ("DÉC", "Décomposition", "Élémentaire", "Corruption, acide, ténèbres"),
]

# --- Rayon d'En : coût DI cumulé + aura/tour ------------------------------- #
RAYON_EN = [
    ("1 m", 0, 5), ("2 m", 10, 5), ("3 m", 10, 5), ("5 m", 10, 10),
    ("10 m", 20, 15), ("20 m", 20, 20), ("30 m", 20, 30), ("50 m", 20, 40),
    ("100 m", 50, 60), ("200 m", 50, 80), ("300 m", 50, 100), ("500 m", 50, 150),
    ("1 km", 100, 200), ("2 km", 100, 300),
]

# --- Caractéristiques (12) ------------------------------------------------- #
CARACS = [
    ("FOR", "Force"), ("DEX", "Dextérité"), ("AGI", "Agilité"),
    ("END", "Endurance"), ("PER", "Perception"), ("PRÉ", "Présence"),
    ("VOL", "Volonté"), ("LOG", "Logique"), ("INS", "Instinct"),
    ("ÉRU", "Érudition"), ("IMA", "Imagination"), ("CHA", "Charisme"),
]
CARAC_DEFAUT = 5  # valeur d'exemple

# --- Compétences par champ : (nom, carac liée) ----------------------------- #
COMPETENCES = {
    "Martial": [
        ("Initiative", "AGI"), ("Armes de mêlée", "DEX"), ("Armes de jet", "DEX"),
        ("Archerie", "DEX"), ("Armes à feu", "DEX"), ("Parade", "DEX"),
        ("Esquive", "AGI"),
    ],
    "Athlétique": [
        ("Course", "AGI"), ("Saut", "AGI"), ("Natation", "AGI"),
        ("Escalade", "FOR"), ("Équilibre", "AGI"), ("Contorsion", "AGI"),
        ("Chute", "AGI"), ("Lutte", "FOR"),
        ("Prouesse de Force", "FOR"), ("Apnée", "END"),
    ],
    "Vital": [
        ("Constitution", "END"), ("Rusticité", "END"),
        ("Ténacité", "END"), ("Impassibilité", "END"),
        ("Courage", "VOL"), ("Sang-froid", "VOL"),
        ("Fermeté", "VOL"), ("Constance", "VOL"),
    ],
    "Social": [
        ("Persuasion", "CHA"), ("Rhétorique", "LOG"), ("Négociation", "CHA"),
        ("Tromperie", "CHA"), ("Séduction", "PRÉ"), ("Intimidation", "PRÉ"),
        ("Interrogatoire", "CHA"), ("Commandement", "PRÉ"), ("Pédagogie", "CHA"),
        ("Relation", "CHA"), ("Étiquette", "ÉRU"), ("Style", "PRÉ"),
    ],
    "Intellectuel": [
        ("Histoire", "ÉRU"), ("Archéologie", "ÉRU"), ("Géographie", "ÉRU"),
        ("Cartographie", "LOG"), ("Économie", "ÉRU"), ("Estimation", "ÉRU"),
        ("Politique", "ÉRU"), ("Droit", "ÉRU"), ("Langues", "ÉRU"),
        ("Astronomie", "ÉRU"), ("Architecture", "ÉRU"), ("Géologie", "ÉRU"),
        ("Faune", "ÉRU"), ("Flore", "ÉRU"), ("Météorologie", "INS"),
        ("Mémorisation", "ÉRU"), ("Occultisme", "ÉRU"), ("Théologie", "ÉRU"),
        ("Médecine", "ÉRU"), ("Enquête", "LOG"), ("Psychologie", "ÉRU"),
        ("Perspicacité", "INS"), ("Tactique", "LOG"), ("Jeu d'argent", "LOG"),
        ("Jeux de stratégie", "LOG"),
    ],
    "Sauvage": [
        ("Survie", "INS"), ("Pistage", "INS"), ("Chasse", "INS"),
        ("Pêche", "INS"), ("Dressage", "INS"), ("Élevage", "INS"),
        ("Agriculture", "INS"),
    ],
    "Technique": [
        ("Mécanique", "LOG"), ("Électronique", "LOG"), ("Programmation", "LOG"),
        ("Cybersécurité", "LOG"), ("Réseaux & télécoms", "LOG"),
        ("Cryptographie", "LOG"), ("Chimie", "LOG"), ("Pharmacologie", "LOG"),
        ("Explosifs", "LOG"), ("Conduite", "AGI"), ("Navigation", "AGI"),
        ("Pilotage", "AGI"), ("Équitation", "AGI"),
    ],
    "Furtif": [
        ("Discrétion", "AGI"), ("Camouflage", "DEX"), ("Filature", "PER"),
        ("Escamotage", "DEX"), ("Crochetage", "DEX"), ("Pièges", "DEX"),
        ("Déguisement", "DEX"), ("Contrefaçon", "DEX"),
    ],
    "Créatif": [
        ("Forge / Métallurgie", "DEX"), ("Travail du bois", "DEX"),
        ("Travail de la pierre", "DEX"), ("Textile", "DEX"),
        ("Travail du cuir", "DEX"), ("Poterie & verre", "DEX"),
        ("Joaillerie", "DEX"), ("Cuisine", "DEX"),
        ("Brasserie & distillation", "DEX"), ("Parfumerie", "DEX"),
        ("Calligraphie", "DEX"), ("Peinture & dessin", "IMA"),
        ("Écriture", "IMA"), ("Conte / Narration", "CHA"),
        ("Musique & chant", "IMA"), ("Danse", "AGI"),
        ("Théâtre", "CHA"), ("Marionnettes", "DEX"),
    ],
}

# --- Sens (champ Sensoriel) : nom + carac (toujours PER) -------------------- #
SENS = [
    "Observation / Vue", "Ouïe", "Odorat", "Goût", "Toucher",
    "Chémoréception", "Chronoception", "Équilibrioception", "Intéroception",
    "Magnétoréception", "Nociception", "Proprioception", "Thermoception",
]


# =========================================================================== #
# CONSTRUCTION DU CLASSEUR
# =========================================================================== #
wb = openpyxl.Workbook()
ws_fiche = wb.active
ws_fiche.title = "Fiche"
ws_bar = wb.create_sheet("Barèmes")
ws_ref = wb.create_sheet("Référence")


# --------------------------------------------------------------------------- #
# FEUILLE « Barèmes »  (construite en premier : la Fiche y pointe)
# --------------------------------------------------------------------------- #
# On mémorise les plages nommées (références A1 absolues) pour les formules.
BAR = {}  # clé -> "Barèmes!$X$r1:$Y$r2"

def bar_block_title(row, c1, c2, text):
    section_title(ws_bar, row, c1, c2, text)

r = 1
cell(ws_bar, "A1", "BARÈMES — tables de correspondance (feuille de calcul, ne pas modifier)",
     font=FONT_LABEL, fill=FILL_PARCHEMIN, align=AL_L, border=False)
ws_bar.merge_cells("A1:H1")
r = 3

# --- Table MOD carac (cols A,B) -------------------------------------------- #
section_title(ws_bar, r, 1, 2, "MOD carac (valeur 0–30 → modificateur)")
r += 1
header_cell(ws_bar, f"A{r}", "Valeur")
header_cell(ws_bar, f"B{r}", "MOD")
r += 1
mod_first = r
for v, m in MOD_TABLE:
    inp_fill = FILL_ALT1 if (v % 2 == 0) else FILL_ALT2
    cell(ws_bar, f"A{r}", v, align=AL_C, fill=inp_fill)
    cell(ws_bar, f"B{r}", m, align=AL_C, fill=inp_fill)
    r += 1
mod_last = r - 1
BAR["MOD"] = f"Barèmes!$A${mod_first}:$B${mod_last}"
r += 1

# --- Plafond carac par Éclat (cols A,B) ------------------------------------ #
section_title(ws_bar, r, 1, 2, "Plafond carac par Éclat")
r += 1
header_cell(ws_bar, f"A{r}", "Éclat ≥")
header_cell(ws_bar, f"B{r}", "Plafond")
r += 1
plaf_first = r
for e, p in PLAFOND_ECLAT:
    cell(ws_bar, f"A{r}", e, align=AL_C, fill=FILL_PARCHEMIN)
    cell(ws_bar, f"B{r}", p, align=AL_C, fill=FILL_PARCHEMIN)
    r += 1
plaf_last = r - 1
BAR["PLAFOND"] = f"Barèmes!$A${plaf_first}:$B${plaf_last}"
r += 1

# --- Création par Éclat (cols A..D) ---------------------------------------- #
section_title(ws_bar, r, 1, 4, "Création par Éclat (moyenne / plafond / budget)")
r += 1
header_cell(ws_bar, f"A{r}", "Éclat ≥")
header_cell(ws_bar, f"B{r}", "Moyenne")
header_cell(ws_bar, f"C{r}", "Plafond")
header_cell(ws_bar, f"D{r}", "Budget")
r += 1
crea_first = r
for e, moy, plaf, bud in CREATION_ECLAT:
    cell(ws_bar, f"A{r}", e, align=AL_C, fill=FILL_PARCHEMIN)
    cell(ws_bar, f"B{r}", moy, align=AL_C, fill=FILL_PARCHEMIN)
    cell(ws_bar, f"C{r}", plaf, align=AL_C, fill=FILL_PARCHEMIN)
    cell(ws_bar, f"D{r}", bud, align=AL_C, fill=FILL_PARCHEMIN)
    r += 1
crea_last = r - 1
BAR["CREATION"] = f"Barèmes!$A${crea_first}:$D${crea_last}"
r += 1

# --- Indice de poids (cols A..D : FOR, Légère, Inter, Lourde) --------------- #
section_title(ws_bar, r, 1, 4, "Indice de poids (FOR → Légère / Inter / Lourde)")
r += 1
header_cell(ws_bar, f"A{r}", "FOR")
header_cell(ws_bar, f"B{r}", "Légère")
header_cell(ws_bar, f"C{r}", "Intermédiaire")
header_cell(ws_bar, f"D{r}", "Lourde")
r += 1
poids_first = r
for v in range(31):
    fill = FILL_ALT1 if (v % 2 == 0) else FILL_ALT2
    leg, inter, lourd = POIDS[v]
    cell(ws_bar, f"A{r}", v, align=AL_C, fill=fill)
    cell(ws_bar, f"B{r}", leg, align=AL_C, fill=fill)
    cell(ws_bar, f"C{r}", inter, align=AL_C, fill=fill)
    cell(ws_bar, f"D{r}", lourd, align=AL_C, fill=fill)
    r += 1
poids_last = r - 1
BAR["POIDS"] = f"Barèmes!$A${poids_first}:$D${poids_last}"
r += 1

# --- Mouvement (cols A..D : AGI, ...) -------------------------------------- #
section_title(ws_bar, r, 1, 4, "Mouvement (AGI → Légère / Inter / Lourde, par round)")
r += 1
header_cell(ws_bar, f"A{r}", "AGI")
header_cell(ws_bar, f"B{r}", "Légère")
header_cell(ws_bar, f"C{r}", "Intermédiaire")
header_cell(ws_bar, f"D{r}", "Lourde")
r += 1
mouv_first = r
for v in range(31):
    fill = FILL_ALT1 if (v % 2 == 0) else FILL_ALT2
    leg, inter, lourd = MOUVEMENT[v]
    cell(ws_bar, f"A{r}", v, align=AL_C, fill=fill)
    cell(ws_bar, f"B{r}", leg, align=AL_C, fill=fill)
    cell(ws_bar, f"C{r}", inter, align=AL_C, fill=fill)
    cell(ws_bar, f"D{r}", lourd, align=AL_C, fill=fill)
    r += 1
mouv_last = r - 1
BAR["MOUVEMENT"] = f"Barèmes!$A${mouv_first}:$D${mouv_last}"
r += 1

# --- Apnée (cols A..D : END, ...) ------------------------------------------ #
section_title(ws_bar, r, 1, 4, "Apnée (END → Légère / Inter / Lourde)")
r += 1
header_cell(ws_bar, f"A{r}", "END")
header_cell(ws_bar, f"B{r}", "Légère")
header_cell(ws_bar, f"C{r}", "Intermédiaire")
header_cell(ws_bar, f"D{r}", "Lourde")
r += 1
apnee_first = r
for v in range(31):
    fill = FILL_ALT1 if (v % 2 == 0) else FILL_ALT2
    leg, inter, lourd = APNEE[v]
    cell(ws_bar, f"A{r}", v, align=AL_C, fill=fill)
    cell(ws_bar, f"B{r}", leg, align=AL_C, fill=fill)
    cell(ws_bar, f"C{r}", inter, align=AL_C, fill=fill)
    cell(ws_bar, f"D{r}", lourd, align=AL_C, fill=fill)
    r += 1
apnee_last = r - 1
BAR["APNEE"] = f"Barèmes!$A${apnee_first}:$D${apnee_last}"
r += 1

# --- Affinités par archétype (cols A..H) ----------------------------------- #
# A=archétype, B=Ren, C=Émi, D=Tra, E=Man, F=Conj, G=Spé sans, H=Spé avec
section_title(ws_bar, r, 1, 8, "Affinités par archétype (%)  —  G = Spé sans Spécialiste, H = Spé avec")
r += 1
for col, txt in zip("ABCDEFGH",
                    ["Archétype", "Renforcement", "Émission", "Transmutation",
                     "Manipulation", "Conjuration", "Spé (sans)", "Spé (avec)"]):
    header_cell(ws_bar, f"{col}{r}", txt)
r += 1
arch_first = r
for i, (name, ren, emi, tra, man, conj, spe0, spe1) in enumerate(ARCHETYPES):
    fill = FILL_ALT1 if (i % 2 == 0) else FILL_ALT2
    cell(ws_bar, f"A{r}", name, align=AL_L, fill=fill)
    for col, val in zip("BCDEFGH", [ren, emi, tra, man, conj, spe0, spe1]):
        cell(ws_bar, f"{col}{r}", val, align=AL_C, fill=fill)
    r += 1
arch_last = r - 1
BAR["ARCH"] = f"Barèmes!$A${arch_first}:$H${arch_last}"
r += 1

# --- Coûts DI des techniques par palier (cumulés) -------------------------- #
# A=technique, B=Aucun, C=Basique, D=Avancé, E=Expert, F=Maître
section_title(ws_bar, r, 1, 6, "Coûts DI des techniques (cumulés par palier)")
r += 1
for col, txt in zip("ABCDEF",
                    ["Technique", "Aucun", "Basique", "Avancé", "Expert", "Maître"]):
    header_cell(ws_bar, f"{col}{r}", txt)
r += 1
techp_first = r
for i, (tech, d) in enumerate(TECH_PALIERS.items()):
    fill = FILL_ALT1 if (i % 2 == 0) else FILL_ALT2
    cell(ws_bar, f"A{r}", tech, align=AL_L, fill=fill)
    for col, pal in zip("BCDEF", ["Aucun", "Basique", "Avancé", "Expert", "Maître"]):
        cell(ws_bar, f"{col}{r}", d[pal], align=AL_C, fill=fill)
    r += 1
techp_last = r - 1
BAR["TECHP"] = f"Barèmes!$A${techp_first}:$F${techp_last}"
r += 1

# --- Base d'UAR par palier d'état à manteau -------------------------------- #
section_title(ws_bar, r, 1, 2, "Base d'UAR par palier (état à manteau actif)")
r += 1
header_cell(ws_bar, f"A{r}", "Palier")
header_cell(ws_bar, f"B{r}", "Base UAR")
r += 1
uar_first = r
for i, (pal, base) in enumerate(UAR_PAR_PALIER.items()):
    fill = FILL_ALT1 if (i % 2 == 0) else FILL_ALT2
    cell(ws_bar, f"A{r}", pal, align=AL_L, fill=fill)
    cell(ws_bar, f"B{r}", base, align=AL_C, fill=fill)
    r += 1
uar_last = r - 1
BAR["UARBASE"] = f"Barèmes!$A${uar_first}:$B${uar_last}"
r += 1

# --- Échelle de difficulté ------------------------------------------------- #
section_title(ws_bar, r, 1, 2, "Échelle de difficulté (palier → seuil)")
r += 1
header_cell(ws_bar, f"A{r}", "Palier")
header_cell(ws_bar, f"B{r}", "Seuil")
r += 1
diff_first = r
for i, (pal, seuil) in enumerate(DIFFICULTE):
    fill = FILL_ALT1 if (i % 2 == 0) else FILL_ALT2
    cell(ws_bar, f"A{r}", pal, align=AL_L, fill=fill)
    cell(ws_bar, f"B{r}", seuil, align=AL_C, fill=fill)
    r += 1
diff_last = r - 1
BAR["DIFF"] = f"Barèmes!$A${diff_first}:$B${diff_last}"
r += 1

# --- Types de dégâts ------------------------------------------------------- #
section_title(ws_bar, r, 1, 4, "Types de dégâts (7)")
r += 1
for col, txt in zip("ABCD", ["Sigle", "Nom", "Famille", "Exemples"]):
    header_cell(ws_bar, f"{col}{r}", txt)
r += 1
typ_first = r
for i, (sig, nom, fam, ex) in enumerate(TYPES_DEGATS):
    fill = FILL_ALT1 if (i % 2 == 0) else FILL_ALT2
    cell(ws_bar, f"A{r}", sig, align=AL_C, fill=fill)
    cell(ws_bar, f"B{r}", nom, align=AL_L, fill=fill)
    cell(ws_bar, f"C{r}", fam, align=AL_L, fill=fill)
    cell(ws_bar, f"D{r}", ex, align=AL_L, fill=fill)
    r += 1
typ_last = r - 1
BAR["TYPES"] = f"Barèmes!$A${typ_first}:$D${typ_last}"
r += 1

# --- Rayon d'En ------------------------------------------------------------ #
section_title(ws_bar, r, 1, 3, "Rayon d'En (rayon → coût DI cumulé + aura/tour)")
r += 1
for col, txt in zip("ABC", ["Rayon", "Coût DI cumulé", "Aura/tour"]):
    header_cell(ws_bar, f"{col}{r}", txt)
r += 1
en_first = r
cum = 0
for i, (rayon, ddi, aura) in enumerate(RAYON_EN):
    cum += ddi
    fill = FILL_ALT1 if (i % 2 == 0) else FILL_ALT2
    cell(ws_bar, f"A{r}", rayon, align=AL_C, fill=fill)
    cell(ws_bar, f"B{r}", cum, align=AL_C, fill=fill)
    cell(ws_bar, f"C{r}", aura, align=AL_C, fill=fill)
    r += 1
en_last = r - 1
BAR["EN"] = f"Barèmes!$A${en_first}:$C${en_last}"
r += 1

# Largeurs Barèmes
for col, w in zip("ABCDEFGH", [26, 16, 16, 16, 14, 14, 12, 12]):
    ws_bar.column_dimensions[col].width = w
ws_bar.sheet_state = "hidden"


# --------------------------------------------------------------------------- #
# Listes de validation (réutilisent les plages de Barèmes / listes inline)
# --------------------------------------------------------------------------- #
def add_list(ws, rng, formula1):
    dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
    dv.add(rng)
    ws.add_data_validation(dv)
    return dv

PALIER_LIST = '"Aucun,Basique,Avancé,Expert,Maître"'
OUINON_LIST = '"Oui,Non"'
CLASSE_LIST = '"Combattant"'
TAILLE_LIST = ('"Microscopique,Infime,Minuscule,Très petite,Petite,Moyenne,'
               'Grande,Très grande,Gigantesque,Colossale,Titanesque"')
TYPE_LIST = '"CON,TRA,PER,FEU,FRO,ÉLE,DÉC"'
# Bonus de classe (Combattant) : choix d'arme d'attaque et de défense
ARME_BONUS_LIST = '"Armes de mêlée,Armes de jet,Archerie,Armes à feu"'
DEF_BONUS_LIST  = '"Parade,Esquive"'
ARCH_LIST_RNG = f"={BAR['ARCH'].split(':')[0]}:$A${arch_last}"  # colonne A des archétypes


# =========================================================================== #
# FEUILLE « Fiche »
# =========================================================================== #
F = ws_fiche
# largeurs
widths = {"A": 22, "B": 12, "C": 10, "D": 14, "E": 3,
          "F": 22, "G": 12, "H": 10, "I": 14, "J": 3,
          "K": 22, "L": 12, "M": 10, "N": 14}
for col, w in widths.items():
    F.column_dimensions[col].width = w

# Repère des coordonnées de carac (pour pointer dessus depuis combat/compétences)
# Renseigné lors de la construction du bloc Caractéristiques.
CARAC_VAL = {}  # sigle -> coord valeur (ex "B12")
CARAC_MOD = {}  # sigle -> coord MOD   (ex "C12")

row = 1
# ---- Titre principal ------------------------------------------------------ #
F.merge_cells(f"A{row}:N{row}")
cell(F, f"A{row}", "FICHE DE PERSONNAGE — Hunter × Hunter (JDR)",
     font=FONT_TITRE, fill=FILL_TITRE, align=AL_L)
for cc in range(1, 15):
    F[f"{get_column_letter(cc)}{row}"].fill = FILL_TITRE
F.row_dimensions[row].height = 28
row += 1

# ---- Légende -------------------------------------------------------------- #
F.merge_cells(f"A{row}:N{row}")
cell(F, f"A{row}",
     "Légende : cellules BLANCHES = à remplir (saisie) ;  cellules BEIGES = calculées automatiquement (ne pas écrire dedans).",
     font=FONT_SMALL, fill=FILL_LEGENDE, align=AL_L)
for cc in range(1, 15):
    F[f"{get_column_letter(cc)}{row}"].fill = FILL_LEGENDE
row += 2

# =========================================================================== #
# 1. IDENTITÉ
# =========================================================================== #
section_title(F, row, 1, 14, "1 · Identité")
row += 1
id_row = row

# Ligne libellés/valeurs sur 2 sous-lignes par bloc de colonnes.
# Colonne A/B : Nom, Classe, Éclat, PF total, DI total, PV courants (saisies)
# Colonne C/D etc : dérivés.
label(F, f"A{row}", "Nom")
NOM = inp(F, f"B{row}", "Gon", align=AL_L)
F.merge_cells(f"B{row}:D{row}")
label(F, f"F{row}", "Classe")
CLASSE = inp(F, f"G{row}", "Combattant", align=AL_L)
F.merge_cells(f"G{row}:I{row}")
add_list(F, f"G{row}", CLASSE_LIST)
row += 1

label(F, f"A{row}", "Éclat")
ECLAT = inp(F, f"B{row}", 10)
label(F, f"C{row}", "Palier d'Éclat")
# Palier = 5*ROUNDDOWN(Éclat/5;0)
PALIER_ECLAT = calc(F, f"D{row}", f"=5*ROUNDDOWN({ECLAT.coordinate}/5,0)")
label(F, f"F{row}", "Plafond carac")
# VLOOKUP(Éclat ; PLAFOND ; 2 ; VRAI)  (table par paliers, recherche approchée)
PLAF_CARAC = calc(F, f"G{row}", f"=VLOOKUP({ECLAT.coordinate},{BAR['PLAFOND']},2,TRUE)")
row += 1

label(F, f"A{row}", "PF total")
PF = inp(F, f"B{row}", 100)
label(F, f"C{row}", "Niveau")
NIVEAU = calc(F, f"D{row}", f"=ROUNDUP({PF.coordinate}/100,0)")
label(F, f"F{row}", "PV max")
PVMAX = calc(F, f"G{row}", f"=100*{NIVEAU.coordinate}")  # Combattant : 100/niveau
row += 1

label(F, f"A{row}", "DI total")
DI = inp(F, f"B{row}", 0)
label(F, f"C{row}", "Prestige")
PRESTIGE = calc(F, f"D{row}", f"=ROUNDUP({DI.coordinate}/100,0)")
label(F, f"F{row}", "PV courants")
PVCUR = inp(F, f"G{row}", 100)
row += 1

# Bonus de classe (Combattant) : deux choix d'office (+10 chacun).
label(F, f"A{row}", "Arme (bonus de classe)")
ARME_BONUS = inp(F, f"B{row}", "Armes de mêlée", align=AL_L)
F.merge_cells(f"B{row}:D{row}")
add_list(F, f"B{row}", ARME_BONUS_LIST)
label(F, f"F{row}", "Défense (bonus de classe)")
DEF_BONUS = inp(F, f"G{row}", "Esquive", align=AL_L)
F.merge_cells(f"G{row}:I{row}")
add_list(F, f"G{row}", DEF_BONUS_LIST)
row += 1
cell(F, f"A{row}",
     "Bonus de classe du Combattant (+10 d'office) : Initiative, Lutte, Impassibilité, "
     "l'arme choisie ci-dessus et la défense choisie ci-dessus. Reportés dans la colonne Bonus des compétences.",
     font=FONT_SMALL, fill=FILL_PARCHEMIN, align=AL_LW, border=False)
F.merge_cells(f"A{row}:N{row}")
F.row_dimensions[row].height = 24
row += 2

# =========================================================================== #
# 1 bis. PLAFONDS DE FORMATION (Combattant)
# =========================================================================== #
# Pour chaque champ : plafond en % (constante) et PF max dérivé =
# ROUNDDOWN(plafond% × PF total / 100, 0), pointant vers la cellule PF total.
section_title(F, row, 1, 14, "1 bis · Plafonds de formation (Combattant)")
row += 1

PLAFONDS_FORMATION = [
    ("Martial", 40), ("Vital", 30), ("Athlétique", 30), ("Sensoriel", 20),
    ("Furtif", 20), ("Sauvage", 20), ("Social", 10), ("Intellectuel", 10),
    ("Technique", 10), ("Créatif", 10),
]

# Trois blocs de colonnes : (Champ | Plafond % | PF max) en A.., F.., K..
PLAF_GROUPS = [("A", "B", "C"), ("F", "G", "H"), ("K", "L", "M")]
PF_REF = f"${PF.coordinate[0]}${PF.row}"  # référence absolue vers PF total

# En-têtes pour chaque bloc
for cChamp, cPct, cMax in PLAF_GROUPS:
    header_cell(F, f"{cChamp}{row}", "Champ")
    header_cell(F, f"{cPct}{row}", "Plafond %")
    header_cell(F, f"{cMax}{row}", "PF max")
row += 1

# 10 champs répartis sur 3 colonnes (4 + 3 + 3)
plaf_layout = [PLAFONDS_FORMATION[0:4], PLAFONDS_FORMATION[4:7], PLAFONDS_FORMATION[7:10]]
plaf_top = row
plaf_bottoms = []
for gi, group in enumerate(plaf_layout):
    cChamp, cPct, cMax = PLAF_GROUPS[gi]
    rr = plaf_top
    for champ, pct in group:
        label(F, f"{cChamp}{rr}", champ)
        # % = constante (texte/calcul), affiché avec le signe %
        calc(F, f"{cPct}{rr}", pct, align=AL_C, fmt="0\\%")
        # PF max = ROUNDDOWN(plafond% × PF total / 100, 0) → formule vers PF total
        calc(F, f"{cMax}{rr}", f"=ROUNDDOWN({pct}*{PF_REF}/100,0)", align=AL_C)
        rr += 1
    plaf_bottoms.append(rr)
row = max(plaf_bottoms)

# Note sur les arts martiaux / formations
cell(F, f"A{row}",
     "Arts martiaux et formations comptent dans le champ Martial. "
     "Le plafond s'applique aux PF totaux, et s'élève donc avec le niveau.",
     font=FONT_SMALL, fill=FILL_PARCHEMIN, align=AL_LW, border=False)
F.merge_cells(f"A{row}:N{row}")
F.row_dimensions[row].height = 24
row += 2

# =========================================================================== #
# 2. CARACTÉRISTIQUES
# =========================================================================== #
section_title(F, row, 1, 14, "2 · Caractéristiques  (saisie 0–30, ≤ plafond)")
row += 1
# En-têtes (deux colonnes de 6 : physiques à gauche, mentales à droite)
header_cell(F, f"A{row}", "Physique")
header_cell(F, f"B{row}", "Valeur")
header_cell(F, f"C{row}", "MOD")
header_cell(F, f"D{row}", "Plafond")
header_cell(F, f"F{row}", "Mentale")
header_cell(F, f"G{row}", "Valeur")
header_cell(F, f"H{row}", "MOD")
header_cell(F, f"I{row}", "Plafond")
row += 1

phys = CARACS[:6]
ment = CARACS[6:]
for i in range(6):
    # physique (cols A,B,C,D)
    sig, nom = phys[i]
    label(F, f"A{row}", f"{sig} — {nom}")
    vcell = inp(F, f"B{row}", CARAC_DEFAUT)
    mcell = calc(F, f"C{row}", f"=VLOOKUP({vcell.coordinate},{BAR['MOD']},2,FALSE)")
    calc(F, f"D{row}", f"=${PLAF_CARAC.coordinate[0]}${PLAF_CARAC.row}")
    CARAC_VAL[sig] = vcell.coordinate
    CARAC_MOD[sig] = mcell.coordinate
    # mentale (cols F,G,H,I)
    sig2, nom2 = ment[i]
    label(F, f"F{row}", f"{sig2} — {nom2}")
    vcell2 = inp(F, f"G{row}", CARAC_DEFAUT)
    mcell2 = calc(F, f"H{row}", f"=VLOOKUP({vcell2.coordinate},{BAR['MOD']},2,FALSE)")
    calc(F, f"I{row}", f"=${PLAF_CARAC.coordinate[0]}${PLAF_CARAC.row}")
    CARAC_VAL[sig2] = vcell2.coordinate
    CARAC_MOD[sig2] = mcell2.coordinate
    row += 1

# Total des caracs
label(F, f"A{row}", "Total caractéristiques")
val_cols_phys = [CARAC_VAL[s] for s, _ in phys]
val_cols_ment = [CARAC_VAL[s] for s, _ in ment]
calc(F, f"B{row}",
     "=SUM(" + ",".join(val_cols_phys + val_cols_ment) + ")")
cell(F, f"C{row}", "Repère : 5 moyenne · 9 max humain · 19 sommet · 30 au-delà",
     font=FONT_SMALL, fill=FILL_PARCHEMIN, align=AL_L, border=False)
F.merge_cells(f"C{row}:I{row}")
row += 2

# =========================================================================== #
# 3. CAPACITÉS PHYSIQUES
# =========================================================================== #
section_title(F, row, 1, 14, "3 · Capacités physiques")
row += 1

# Actions actives / tour
label(F, f"A{row}", "Actions actives / tour")
ACTIONS = calc(F, f"B{row}",
     f"=MAX(1,MIN(10,ROUNDDOWN(({CARAC_MOD['AGI']}+{CARAC_MOD['DEX']})/10,0)))")
label(F, f"D{row}", "Taille")
TAILLE = inp(F, f"E{row}", "Moyenne", align=AL_L)
F.merge_cells(f"E{row}:G{row}")
add_list(F, f"E{row}", TAILLE_LIST)
label(F, f"H{row}", "Espace / Allonge")
cell(F, f"I{row}", "1 m / 1 m (Moyenne)", font=FONT_CALC, fill=FILL_PARCHEMIN, align=AL_C)
F.merge_cells(f"I{row}:J{row}")
row += 1

# En-tête L / I / Lo
header_cell(F, f"A{row}", "Capacité")
header_cell(F, f"B{row}", "Légère")
header_cell(F, f"C{row}", "Intermédiaire")
header_cell(F, f"D{row}", "Lourde")
header_cell(F, f"F{row}", "Capacité")
header_cell(F, f"G{row}", "Légère")
header_cell(F, f"H{row}", "Intermédiaire")
header_cell(F, f"I{row}", "Lourde")
row += 1

# Mouvement (AGI)  — colonnes gauche
label(F, f"A{row}", "Mouvement")
for ci, col in enumerate("BCD", start=2):
    calc(F, f"{col}{row}",
         f"=VLOOKUP({CARAC_VAL['AGI']},{BAR['MOUVEMENT']},{ci},FALSE)")
# Indice de poids (FOR) — colonnes droite
label(F, f"F{row}", "Indice de poids")
for ci, col in enumerate("GHI", start=2):
    calc(F, f"{col}{row}",
         f"=VLOOKUP({CARAC_VAL['FOR']},{BAR['POIDS']},{ci},FALSE)")
row += 1

# Apnée (END) gauche ; Tirer/pousser note droite
label(F, f"A{row}", "Apnée")
for ci, col in enumerate("BCD", start=2):
    calc(F, f"{col}{row}",
         f"=VLOOKUP({CARAC_VAL['END']},{BAR['APNEE']},{ci},FALSE)")
label(F, f"F{row}", "Tirer / pousser")
cell(F, f"G{row}", "= indice de poids × 10 (chaque colonne)",
     font=FONT_SMALL, fill=FILL_PARCHEMIN, align=AL_L, border=True)
F.merge_cells(f"G{row}:I{row}")
row += 1

# Suffocation (END)
label(F, f"A{row}", "Suffocation")
SUFFO = calc(F, f"B{row}", f"={CARAC_VAL['END']}", align=AL_C)
cell(F, f"C{row}", "rounds avant inconscience ; mort à 2 × END",
     font=FONT_SMALL, fill=FILL_PARCHEMIN, align=AL_L, border=True)
F.merge_cells(f"C{row}:D{row}")
label(F, f"F{row}", "Mort (suffocation)")
calc(F, f"G{row}", f"=2*{CARAC_VAL['END']}", align=AL_C)
cell(F, f"H{row}", "rounds (pleine respiration → mort)",
     font=FONT_SMALL, fill=FILL_PARCHEMIN, align=AL_L, border=True)
F.merge_cells(f"H{row}:I{row}")
row += 2

# =========================================================================== #
# 4. COMBAT
# =========================================================================== #
# Les compétences de combat (Initiative, Armes de mêlée/jet, Archerie, Armes à
# feu, Parade, Esquive, Lutte) sont des compétences génériques ordinaires : elles
# vivent dans la section « 7 · Compétences » (champs Martial / Athlétique) et ne
# sont PAS dupliquées ici. Le bloc Combat ne garde que ce qui lui est propre :
# rythme d'action, PV, armure (TA) et arme équipée.
section_title(F, row, 1, 14, "4 · Combat")
row += 1

# Rythme d'action + PV
label(F, f"A{row}", "Actions actives / tour")
calc(F, f"B{row}", f"=${ACTIONS.coordinate[0]}${ACTIONS.row}")
label(F, f"D{row}", "PV max")
calc(F, f"E{row}", f"=${PVMAX.coordinate[0]}${PVMAX.row}")
label(F, f"G{row}", "PV courants")
calc(F, f"H{row}", f"=${PVCUR.coordinate[0]}${PVCUR.row}")
row += 2

# TA (armure) : 7 cases de saisie, une par type de dégâts
label(F, f"A{row}", "TA (armure)")
F.merge_cells(f"A{row}:A{row+1}")
ta_types = ["CON", "TRA", "PER", "FEU", "FRO", "ÉLE", "DÉC"]
for i, t in enumerate(ta_types):
    col = get_column_letter(2 + i)   # B..H
    header_cell(F, f"{col}{row}", t)
    inp(F, f"{col}{row+1}", 0)
row += 2

# Renvoi vers les compétences pour les jets de combat
cell(F, f"A{row}",
     "Jets de combat — Initiative, attaque avec l'arme (Armes de mêlée / jet / "
     "Archerie / feu), Parade, Esquive, Lutte : voir la section Compétences.",
     font=FONT_SMALL, fill=FILL_PARCHEMIN, align=AL_LW, border=False)
F.merge_cells(f"A{row}:N{row}")
row += 2

# Arme équipée
label(F, f"A{row}", "Arme équipée")
section_sub = row
header_cell(F, f"B{row}", "Nom")
header_cell(F, f"C{row}", "Dégâts faciaux")
header_cell(F, f"D{row}", "Mult. N (0–3)")
header_cell(F, f"F{row}", "Finesse")
header_cell(F, f"G{row}", "Type")
header_cell(F, f"H{row}", "Dégâts de base")
row += 1
label(F, f"A{row}", "")
ARME_NOM = inp(F, f"B{row}", "Mains nues", align=AL_L)
ARME_FAC = inp(F, f"C{row}", 20)
ARME_N = inp(F, f"D{row}", 1)
ARME_FIN = inp(F, f"F{row}", "Non")
add_list(F, f"F{row}", OUINON_LIST)
ARME_TYPE = inp(F, f"G{row}", "CON")
add_list(F, f"G{row}", TYPE_LIST)
# Dégâts de base = faciaux + N × SI(Finesse ; MOD DEX ; MOD FOR)
calc(F, f"H{row}",
     f'={ARME_FAC.coordinate}+{ARME_N.coordinate}*IF({ARME_FIN.coordinate}="Oui",{CARAC_MOD["DEX"]},{CARAC_MOD["FOR"]})')
row += 1
cell(F, f"A{row}", "Dégâts infligés = (Dégâts de base − TA du type) × marge% ; la TA se retire avant le multiplicateur.",
     font=FONT_SMALL, fill=FILL_PARCHEMIN, align=AL_L, border=False)
F.merge_cells(f"A{row}:N{row}")
row += 2

# =========================================================================== #
# 5. AURA & NEN
# =========================================================================== #
section_title(F, row, 1, 14, "5 · Aura & Nen")
row += 1

# Archétype + Spécialiste -> affinités
label(F, f"A{row}", "Archétype")
ARCH = inp(F, f"B{row}", "Renforceur", align=AL_L)
F.merge_cells(f"B{row}:C{row}")
add_list(F, f"B{row}", ARCH_LIST_RNG)
label(F, f"D{row}", "Avantage Spécialiste")
SPE = inp(F, f"F{row}", "Non")
add_list(F, f"F{row}", OUINON_LIST)
label(F, f"H{row}", "DI total")
calc(F, f"I{row}", f"=${DI.coordinate[0]}${DI.row}")
row += 1

# 6 affinités
header_cell(F, f"A{row}", "Affinité")
header_cell(F, f"B{row}", "Renf.")
header_cell(F, f"C{row}", "Émis.")
header_cell(F, f"D{row}", "Transm.")
header_cell(F, f"F{row}", "Manip.")
header_cell(F, f"G{row}", "Conj.")
header_cell(F, f"H{row}", "Spé.")
row += 1
label(F, f"A{row}", "Affinité (%)")
# VLOOKUP archétype : col 2=Ren,3=Émi,4=Tra,5=Man,6=Conj,7=Spé sans,8=Spé avec
AFF_CELLS = {}
for col, idx, key in [("B", 2, "Ren"), ("C", 3, "Émi"), ("D", 4, "Tra"),
                      ("F", 5, "Man"), ("G", 6, "Conj")]:
    cc = calc(F, f"{col}{row}",
              f"=VLOOKUP({ARCH.coordinate},{BAR['ARCH']},{idx},FALSE)",
              fmt="0\\%")
    AFF_CELLS[key] = cc.coordinate
# Spé : SI(Spécialiste="Oui" ; col 8 ; col 7)
spe_cc = calc(F, f"H{row}",
     f'=IF({SPE.coordinate}="Oui",VLOOKUP({ARCH.coordinate},{BAR["ARCH"]},8,FALSE),VLOOKUP({ARCH.coordinate},{BAR["ARCH"]},7,FALSE))',
     fmt="0\\%")
AFF_CELLS["Spé"] = spe_cc.coordinate
row += 2

# UAR / RUA / UAM / UAD
header_cell(F, f"A{row}", "Aura")
header_cell(F, f"B{row}", "Base")
header_cell(F, f"C{row}", "Mult.")
header_cell(F, f"D{row}", "Valeur")
header_cell(F, f"F{row}", "Réserve")
header_cell(F, f"G{row}", "Valeur")
row += 1

# UAR = base × (mult + 1)
label(F, f"A{row}", "UAR (par round)")
BASE_UAR = inp(F, f"B{row}", 5)
MULT_UAR = inp(F, f"C{row}", 0)
UAR = calc(F, f"D{row}", f"={BASE_UAR.coordinate}*({MULT_UAR.coordinate}+1)")
# UAM = 100*Prestige + 20*achats
label(F, f"F{row}", "UAM (aura max)")
# rappel prestige + achats
row_uam = row
row += 1

label(F, f"A{row}", "RUA (par heure)")
BASE_RUA = inp(F, f"B{row}", 5)
MULT_RUA = inp(F, f"C{row}", 0)
RUA = calc(F, f"D{row}", f"={BASE_RUA.coordinate}*({MULT_RUA.coordinate}+1)")
# achats UAM
label(F, f"F{row}", "Achats UAM (tranches)")
ACHATS_UAM = inp(F, f"G{row}", 0)
# UAM value placed at row_uam col G
UAM = calc(F, f"G{row_uam}",
     f"=100*${PRESTIGE.coordinate[0]}${PRESTIGE.row}+20*{ACHATS_UAM.coordinate}")
row += 1

label(F, f"A{row}", "Prestige (rappel)")
calc(F, f"B{row}", f"=${PRESTIGE.coordinate[0]}${PRESTIGE.row}")
label(F, f"C{row}", "UAD courant")
UAD = inp(F, f"D{row}", 100)
cell(F, f"F{row}", "UAD ≤ UAM ; remonte de la RUA chaque heure.",
     font=FONT_SMALL, fill=FILL_PARCHEMIN, align=AL_L, border=True)
F.merge_cells(f"F{row}:I{row}")
row += 2

# Conversion DI -> catégorie
label(F, f"A{row}", "Conversion DI → catégorie")
F.merge_cells(f"A{row}:I{row}")
row += 1
header_cell(F, f"A{row}", "Catégorie")
header_cell(F, f"B{row}", "DI investi")
header_cell(F, f"C{row}", "Affinité %")
header_cell(F, f"D{row}", "Points (DX)")
header_cell(F, f"F{row}", "Catégorie")
header_cell(F, f"G{row}", "DI investi")
header_cell(F, f"H{row}", "Affinité %")
header_cell(F, f"I{row}", "Points (DX)")
row += 1
# DR/DE/DT  (gauche)   DM/DC/DS (droite)
conv_left = [("DR (Renforcement)", "Ren"), ("DE (Émission)", "Émi"), ("DT (Transmutation)", "Tra")]
conv_right = [("DM (Manipulation)", "Man"), ("DC (Conjuration)", "Conj"), ("DS (Spécialisation)", "Spé")]
for i in range(3):
    # gauche
    nameL, keyL = conv_left[i]
    label(F, f"A{row}", nameL)
    diL = inp(F, f"B{row}", 0)
    calc(F, f"C{row}", f"={AFF_CELLS[keyL]}", fmt="0\\%")
    calc(F, f"D{row}", f"=ROUNDDOWN({diL.coordinate}*{AFF_CELLS[keyL]}/100,0)")
    # droite
    nameR, keyR = conv_right[i]
    label(F, f"F{row}", nameR)
    diR = inp(F, f"G{row}", 0)
    calc(F, f"H{row}", f"={AFF_CELLS[keyR]}", fmt="0\\%")
    calc(F, f"I{row}", f"=ROUNDDOWN({diR.coordinate}*{AFF_CELLS[keyR]}/100,0)")
    row += 1
row += 1

# =========================================================================== #
# 6. TECHNIQUES DU NEN
# =========================================================================== #
section_title(F, row, 1, 14, "6 · Techniques du Nen")
row += 1

# Initiation + blocs Oui/Non (Hatsu, Gyo, Shu, In)
header_cell(F, f"A{row}", "Technique (bloc)")
header_cell(F, f"B{row}", "Possédée")
header_cell(F, f"C{row}", "Coût DI")
header_cell(F, f"F{row}", "État (palier)")
header_cell(F, f"G{row}", "Palier")
header_cell(F, f"H{row}", "Coût DI cumulé")
row += 1

bloc_start = row
bloc_cells = {}  # nom -> (case coord)
blocs_list = [("Initiation", 40), ("Hatsu", 30), ("Gyo", 20), ("Shu", 20), ("In", 30)]
etats_list = ["Ten", "Zetsu", "Ren", "Ken", "En", "Ko", "Ryu"]

n = max(len(blocs_list), len(etats_list))
etat_cells = {}  # nom -> palier coord
for i in range(n):
    rr = bloc_start + i
    if i < len(blocs_list):
        bname, bcost = blocs_list[i]
        label(F, f"A{rr}", bname)
        bc = inp(F, f"B{rr}", "Non" if bname != "Initiation" else "Oui")
        add_list(F, f"B{rr}", OUINON_LIST)
        calc(F, f"C{rr}", f'=IF({bc.coordinate}="Oui",{bcost},0)')
        bloc_cells[bname] = bc.coordinate
    if i < len(etats_list):
        ename = etats_list[i]
        label(F, f"F{rr}", ename)
        ec = inp(F, f"G{rr}", "Aucun", align=AL_L)
        add_list(F, f"G{rr}", PALIER_LIST)
        # coût cumulé : VLOOKUP technique sur ligne, puis HLOOKUP/INDEX-MATCH palier
        # On utilise INDEX(MOD plage ; MATCH technique ; MATCH palier)
        # TECHP : A=technique, B..F = Aucun..Maître
        # colonne du palier = MATCH(palier ; en-têtes Aucun..Maître)
        techp_rng = BAR['TECHP']
        # en-têtes des paliers : on connaît leur position (Aucun=1 etc. dans B..F)
        # MATCH(ec ; {"Aucun";"Basique";"Avancé";"Expert";"Maître"} ; 0) -> 1..5
        # INDEX(B..F ; MATCH(technique ; A ; 0) ; col)
        calc(F, f"H{rr}",
             f'=INDEX({techp_rng},MATCH("{ename}",{techp_rng.split(":")[0]}:$A${techp_last},0),'
             f'MATCH({ec.coordinate},{{"Aucun","Basique","Avancé","Expert","Maître"}},0)+1)')
        etat_cells[ename] = ec.coordinate

row = bloc_start + n + 1

# Sous-techniques Oui/Non additionnelles : Hatsu déjà ; ajoutons Gyo/Shu/In déjà couverts.
# DI dépensé en techniques (somme)
label(F, f"A{row}", "DI dépensé en techniques")
# somme blocs (C col) + somme états (H col)
bloc_c_cells = [f"C{bloc_start + i}" for i in range(len(blocs_list))]
etat_h_cells = [f"H{bloc_start + i}" for i in range(len(etats_list))]
calc(F, f"B{row}", "=SUM(" + ",".join(bloc_c_cells + etat_h_cells) + ")")
F.merge_cells(f"B{row}:C{row}")
cell(F, f"F{row}",
     "Base d'UAR conférée par l'état à manteau actif (Ren/Ken/Ko/Ryu) : Basique 10 · Avancé 15 · Expert 20 · Maître 25.",
     font=FONT_SMALL, fill=FILL_PARCHEMIN, align=AL_L, border=True)
F.merge_cells(f"F{row}:N{row}")
row += 2

# =========================================================================== #
# 7. COMPÉTENCES
# =========================================================================== #
section_title(F, row, 1, 14, "7 · Compétences  (Bonus = Valeur + MOD de la carac liée)")
row += 1

# On dispose les champs en 3 colonnes de blocs (A.. / F.. / K..)
# Chaque bloc : sous-titre champ, en-tête, lignes (libellé | carac | valeur | bonus)
COL_GROUPS = [
    ("A", "B", "C", "D"),   # bloc gauche
    ("F", "G", "H", "I"),   # bloc milieu
    ("K", "L", "M", "N"),   # bloc droit
]

# --- Bonus de classe (Combattant) appliqués au calcul du Bonus de compétence -- #
# Toujours +10 (fixe) :
CLASS_BONUS_FIXED = {"Initiative", "Lutte", "Impassibilité"}
# Au choix d'arme d'attaque (+10 à celle pointée par la cellule ARME_BONUS) :
CLASS_BONUS_WEAPONS = {"Armes de mêlée", "Armes de jet", "Archerie", "Armes à feu"}
# Au choix de défense (+10 à celle pointée par la cellule DEF_BONUS) :
CLASS_BONUS_DEFENSE = {"Parade", "Esquive"}

def class_bonus_term(skill_name):
    """Renvoie le terme de formule à ajouter à 'Valeur + MOD' pour le bonus de
    classe Combattant, ou '' si la compétence n'en bénéficie pas."""
    if skill_name in CLASS_BONUS_FIXED:
        return "+10"
    if skill_name in CLASS_BONUS_WEAPONS:
        return f'+IF({ARME_BONUS.coordinate}="{skill_name}",10,0)'
    if skill_name in CLASS_BONUS_DEFENSE:
        return f'+IF({DEF_BONUS.coordinate}="{skill_name}",10,0)'
    return ""

def write_field_block(ws, top_row, cols, champ, comps, sensoriel=False):
    """Écrit un champ de compétences dans un bloc de colonnes.
       Renvoie la dernière ligne utilisée."""
    cL, cC, cV, cB = cols
    rr = top_row
    # sous-titre champ
    a = f"{cL}{rr}"
    b = f"{cB}{rr}"
    ws.merge_cells(f"{a}:{b}")
    cell(ws, a, f"Champ {champ}", font=FONT_ENTETE, fill=FILL_SECTION, align=AL_L)
    for cc in (cL, cC, cV, cB):
        ws[f"{cc}{rr}"].fill = FILL_SECTION
        ws[f"{cc}{rr}"].border = BORDER
    rr += 1
    # en-tête
    if sensoriel:
        header_cell(ws, f"{cL}{rr}", "Sens")
        header_cell(ws, f"{cC}{rr}", "Acuité")
        header_cell(ws, f"{cV}{rr}", "Valeur")
        header_cell(ws, f"{cB}{rr}", "Bonus / Malus")
    else:
        header_cell(ws, f"{cL}{rr}", "Compétence")
        header_cell(ws, f"{cC}{rr}", "Carac")
        header_cell(ws, f"{cV}{rr}", "Valeur")
        header_cell(ws, f"{cB}{rr}", "Bonus")
    rr += 1
    for idx, item in enumerate(comps):
        fill = FILL_ALT2 if (idx % 2 == 0) else FILL_ALT1
        if sensoriel:
            sname = item
            cell(ws, f"{cL}{rr}", sname, font=FONT_BASE, fill=fill, align=AL_L)
            acu = inp(ws, f"{cC}{rr}", 10)
            valc = inp(ws, f"{cV}{rr}", 0)
            # Bonus normal = valeur + MOD PER ; mais pour Sensoriel on affiche
            # le malus de perception = (10-acuité)*-20 (et note malus action)
            calc(ws, f"{cB}{rr}",
                 f"=(10-{acu.coordinate})*-20", align=AL_C)
        else:
            sname, csig = item
            cell(ws, f"{cL}{rr}", sname, font=FONT_BASE, fill=fill, align=AL_L)
            cell(ws, f"{cC}{rr}", csig, font=FONT_BASE, fill=fill, align=AL_C)
            valc = inp(ws, f"{cV}{rr}", 0)
            # Bonus = Valeur + MOD (+ bonus de classe Combattant si applicable)
            bonus = class_bonus_term(sname)
            calc(ws, f"{cB}{rr}",
                 f"={valc.coordinate}+{CARAC_MOD[csig]}{bonus}", align=AL_C)
        rr += 1
    return rr

# Ordre des champs et placement en colonnes (équilibrer la hauteur)
# Colonne gauche : Martial, Athlétique, Vital, Furtif
# Colonne milieu : Social, Sauvage, Technique, Créatif
# Colonne droite : Intellectuel, Sensoriel
layout = {
    0: ["Martial", "Athlétique", "Vital", "Furtif"],
    1: ["Social", "Sauvage", "Technique", "Créatif"],
    2: ["Intellectuel", "__SENS__"],
}

base_row = row
col_bottom = {}
for gi, champs in layout.items():
    cols = COL_GROUPS[gi]
    rr = base_row
    for champ in champs:
        if champ == "__SENS__":
            # bloc Sensoriel (avec colonnes Acuité)
            # sous-titre + note
            rr = write_field_block(F, rr, cols, "Sensoriel (13 sens)", SENS, sensoriel=True)
            # note malus d'action
            cL, cC, cV, cB = cols
            F.merge_cells(f"{cL}{rr}:{cB}{rr}")
            cell(F, f"{cL}{rr}",
                 "Malus perception = (10−acuité)×−20 (affiché). Malus d'action = (10−acuité)×−10.",
                 font=FONT_SMALL, fill=FILL_PARCHEMIN, align=AL_LW, border=False)
            rr += 2
        else:
            rr = write_field_block(F, rr, cols, champ, COMPETENCES[champ], sensoriel=False)
            rr += 1  # espace entre champs
    col_bottom[gi] = rr

row = max(col_bottom.values()) + 1

# Note de bas
F.merge_cells(f"A{row}:N{row}")
cell(F, f"A{row}",
     "Total d'un test = d100 + carac liée (MOD) + valeur de compétence, comparé au seuil de difficulté (feuille Référence). "
     "Le Bonus affiché (Valeur + MOD) est ce qui s'ajoute au d100 en jeu.",
     font=FONT_SMALL, fill=FILL_LEGENDE, align=AL_LW, border=False)
F.row_dimensions[row].height = 26

# Figer les volets sous le titre + légende
F.freeze_panes = "A4"


# =========================================================================== #
# FEUILLE « Référence »
# =========================================================================== #
R = ws_ref
for col, w in zip("ABCDEFGH", [22, 16, 16, 16, 16, 16, 16, 16]):
    R.column_dimensions[col].width = w

rr = 1
R.merge_cells(f"A{rr}:H{rr}")
cell(R, f"A{rr}", "RÉFÉRENCE — aide-mémoire (consultatif)",
     font=FONT_TITRE, fill=FILL_TITRE, align=AL_L)
for cc in range(1, 9):
    R[f"{get_column_letter(cc)}{rr}"].fill = FILL_TITRE
R.row_dimensions[rr].height = 26
rr += 2

def ref_table(start, title, headers, rows, ncol):
    section_title(R, start, 1, ncol, title)
    s = start + 1
    for j, h in enumerate(headers):
        header_cell(R, f"{get_column_letter(j+1)}{s}", h)
    s += 1
    for i, rdata in enumerate(rows):
        fill = FILL_ALT1 if (i % 2 == 0) else FILL_ALT2
        for j, val in enumerate(rdata):
            al = AL_L if j == 0 else AL_C
            cell(R, f"{get_column_letter(j+1)}{s}", val, font=FONT_BASE,
                 fill=fill, align=al)
        s += 1
    return s + 1

# Échelle de difficulté
rr = ref_table(rr, "Échelle de difficulté", ["Palier", "Seuil"],
               [[p, s] for p, s in DIFFICULTE], 2)

# Types de dégâts
rr = ref_table(rr, "Types de dégâts (7)", ["Sigle", "Nom", "Famille", "Exemples"],
               [[s, n, f, e] for s, n, f, e in TYPES_DEGATS], 4)

# Seuils de marge
rr = ref_table(rr, "Marge (% de touche)", ["Condition", "Effet"],
               [["≥ 200", "200 % (×2, plafond)"],
                ["10 à 199", "ce % de touche"],
                ["0 à 9", "touche mais 0 dégât"],
                ["< 0", "raté"]], 2)

# Postures
rr = ref_table(rr, "Postures de combat", ["Posture", "Attaque", "Défense"],
               [["Défensive partielle", "−30", "+10"],
                ["Défensive totale", "aucune att.", "+30"],
                ["Offensive partielle", "+10", "−30"],
                ["Offensive totale", "+30", "aucune déf."]], 3)

# Situations
rr = ref_table(rr, "Situations", ["Situation", "Init", "Att", "Esq", "Par", "Phys"],
               [["De dos", "–", "−30", "−80", "−80", "–"],
                ["De flanc", "–", "−10", "−30", "−30", "–"],
                ["En joue", "−50", "−20", "−120", "−120", "−100"],
                ["Espace réduit", "–", "−40", "−40", "−40", "−20"],
                ["Position supérieure", "–", "+20", "–", "–", "–"],
                ["Position inférieure", "–", "−20", "–", "–", "–"],
                ["Surpris", "−120", "–", "–", "–", "–"]], 6)

# États / afflictions (résumé)
rr = ref_table(rr, "États / afflictions (malus)", ["État", "Effet"],
               [["Immobilisé", "−20 att/esq ; +20 contre lui"],
                ["Étourdi", "perd actions, pas de réaction ; +20 contre lui"],
                ["À terre", "−20 att ; +20 mêlée contre / −20 distance contre"],
                ["Paralysie légère", "−20 att/par/init ; −40 esq/phys"],
                ["Paralysie partielle", "−80 att/par/esq ; −30 init ; −60 phys"],
                ["Paralysie totale", "−200 att/par/esq/phys ; −100 init"],
                ["Inconscience", "toute attaque mêlée qui touche = critique"]], 2)

# Manœuvres
rr = ref_table(rr, "Manœuvres (malus à l'attaque)", ["Manœuvre", "Malus / effet"],
               [["Blesser", "0 — seule manœuvre infligeant les dégâts"],
                ["Agripper", "−40 puis lutte → cible Agrippée"],
                ["Mettre à terre", "−40 puis lutte/équilibre → À terre"],
                ["Pousser", "−40 puis lutte/équilibre → repoussé 2 m+"],
                ["Désarmer", "−20 + Viser main (−60 à 1 main / −100 à 2 mains)"],
                ["Maîtriser / Soumettre", "pas de jet d'attaque (lutte opposée)"]], 2)

# Viser
rr = ref_table(rr, "Viser (malus à l'attaque par zone)", ["Zone", "Malus"],
               [["Mollet / Torse", "−10"], ["Abdomen / Bras / Cuisse", "−20"],
                ["Épaule", "−30"], ["Main / Genou / Poignet", "−40"],
                ["Pied", "−50"], ["Tête / Coude / Cœur / Aine", "−60"],
                ["Cou", "−80"], ["Œil", "−100"]], 2)

# Catalogue d'armes (sélection représentative)
ARMES_REF = [
    ["Mains nues", "20", "×1 FOR", "CON", "Indésarmable"],
    ["Dague", "30", "×1", "PER", "Finesse, Dissimulable, jet 10/20 m"],
    ["Épée courte", "40", "×1", "TRA", "Finesse, Dégainage inst."],
    ["Épée longue", "30 (50)", "×2", "TRA", "Polyvalente"],
    ["Épée à 2 mains", "50", "×3", "TRA", "Lourde"],
    ["Sabre", "40 (60)", "×1", "TRA", "Légère, Polyvalente"],
    ["Hache 1 main", "40 (60)", "×2", "TRA", "Polyvalente"],
    ["Masse d'armes", "20", "×2", "CON", "Perce-blindage"],
    ["Lance", "20 (40)", "×2", "PER", "Polyvalente, jet 10/20 m"],
    ["Bâton", "20 (40)", "×1", "CON", "Polyvalente"],
    ["Fouet", "0", "×1", "CON", "Allonge ×3, Désarmement, Saisie"],
    ["Javelot", "60", "×1", "PER", "Fragile ; jet 40/80 m"],
    ["Arc court", "50", "×1", "PER", "tir 30/90 m"],
    ["Arc long", "60", "×1", "PER", "Encombrante ; tir 40/180 m"],
    ["Arbalète légère", "140", "×0", "PER", "Munitions 1"],
    ["Pistolet semi-auto", "145", "×0", "PER", "Dissimulable, Mun. 15"],
    ["Revolver", "155", "×0", "PER", "Dégainage inst., Mun. 6"],
    ["Fusil d'assaut", "175", "×0", "PER", "Rafale, Bruyante, Mun. 30"],
    ["Fusil de précision", "215", "×0", "PER", "Encombr. extrême, Mun. 6"],
    ["Fusil anti-matériel", "240", "×0", "PER", "Perce-blindage, Affût, Mun. 10"],
    ["Grenade à main", "180", "×0", "CON/PER", "Zone 10 m, Usage unique ; jet 30 m"],
    ["Lance-flammes", "155", "×0", "FEU", "Zone cône 10 m, Incendiaire, Mun. 6"],
    ["Griffes", "40", "×1", "TRA", "Finesse, Mains libres"],
    ["Sai / Jutte", "60", "×1", "CON", "Désarmement"],
]
rr = ref_table(rr, "Catalogue d'armes (extrait — chaque arme vaut 100 points)",
               ["Arme", "Dégâts faciaux", "Mod.", "Type", "Propriétés"],
               ARMES_REF, 5)

cell(R, f"A{rr}",
     "Toutes les valeurs proviennent du livre. Feuille purement consultative.",
     font=FONT_SMALL, fill=FILL_PARCHEMIN, align=AL_L, border=False)


# =========================================================================== #
# SAUVEGARDE
# =========================================================================== #
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                   "fiche-personnage-hxh.xlsx")
wb.save(OUT)
print("Écrit :", OUT)
